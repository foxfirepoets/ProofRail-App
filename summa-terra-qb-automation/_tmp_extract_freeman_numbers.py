"""Extract Freeman recon source numbers from Drive sheets/PDFs."""
from __future__ import annotations

import io
import sys
from pathlib import Path

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

sys.stdout.reconfigure(encoding="utf-8")
ENV = Path(r"C:\Users\Heather Workman\Desktop\Ben Projects\Summa Terra Gmail Automation\.env")
OUT = Path(r"C:\Users\Heather Workman\Desktop\Ben Projects\Summa Terra QB Automation\_tmp_freeman_recon_docs")
OUT.mkdir(parents=True, exist_ok=True)


def load_env(path: Path) -> dict:
    vals = {}
    for line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            k, _, v = line.partition("=")
            vals[k.strip()] = v.strip().strip('"').strip("'")
    return vals


def creds():
    env = load_env(ENV)
    c = Credentials(
        token=None,
        refresh_token=env["STONE_REFRESH_TOKEN"],
        token_uri="https://oauth2.googleapis.com/token",
        client_id=env["GOOGLE_CLIENT_ID"],
        client_secret=env["GOOGLE_CLIENT_SECRET"],
    )
    c.refresh(Request())
    return c


def download(drive, file_id: str, dest: Path):
    req = drive.files().get_media(fileId=file_id, supportsAllDrives=True)
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, req)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    dest.write_bytes(fh.getvalue())
    print(f"Downloaded {dest.name} ({dest.stat().st_size} bytes)")


def export_sheet_xlsx(drive, file_id: str, dest: Path):
    # works for both native google sheets and uploaded xlsx if we use get_media for binary
    meta = drive.files().get(fileId=file_id, fields="mimeType,name", supportsAllDrives=True).execute()
    mime = meta["mimeType"]
    print(f"File {meta['name']} mime={mime}")
    if mime == "application/vnd.google-apps.spreadsheet":
        req = drive.files().export_media(
            fileId=file_id,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        req = drive.files().get_media(fileId=file_id, supportsAllDrives=True)
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, req)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    dest.write_bytes(fh.getvalue())
    print(f"Saved {dest} ({dest.stat().st_size})")


def dump_xlsx(path: Path, max_rows=80):
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"\n===== SHEET: {sn} dims={ws.dimensions} =====")
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i > max_rows:
                print(f"... truncated at {max_rows}")
                break
            vals = [("" if v is None else v) for v in row]
            # skip fully empty
            if not any(str(v).strip() for v in vals):
                continue
            print(f"{i:03d}| " + " | ".join(str(v) for v in vals[:12]))


def dump_pdf_text(path: Path, max_chars=6000):
    try:
        from pypdf import PdfReader
    except ImportError:
        try:
            from PyPDF2 import PdfReader
        except ImportError:
            print("No PDF library; skip", path)
            return
    reader = PdfReader(str(path))
    text = []
    for page in reader.pages:
        text.append(page.extract_text() or "")
    full = "\n".join(text)
    print(f"\n===== PDF: {path.name} pages={len(reader.pages)} =====")
    print(full[:max_chars])


def main():
    c = creds()
    drive = build("drive", "v3", credentials=c)
    sheets = build("sheets", "v4", credentials=c)

    # Pricing comparison workbook (uploaded xlsx on Drive)
    pricing_id = "10BZtKURPkghr1iC5gzQwvpKIxdpzxQuG"
    pricing_path = OUT / "7-13-2026_Updated_Pricing_Comparison.xlsx"
    export_sheet_xlsx(drive, pricing_id, pricing_path)
    dump_xlsx(pricing_path, max_rows=120)

    # Also try Sheets API values in case it's a google sheet
    try:
        meta = drive.files().get(fileId=pricing_id, fields="mimeType", supportsAllDrives=True).execute()
        if meta["mimeType"] == "application/vnd.google-apps.spreadsheet":
            vals = sheets.spreadsheets().values().get(spreadsheetId=pricing_id, range="A1:Z100").execute()
            print("\nSheets API rows:", len(vals.get("values", [])))
            for i, row in enumerate(vals.get("values", [])[:80], 1):
                print(f"{i:03d}| " + " | ".join(str(x) for x in row))
    except Exception as e:
        print("Sheets API note:", e)

    # Download key PDFs
    pdfs = {
        "1JUFWTGvh7BG4DY_R0BIRX9feYrfZO6HH": "7-15-2026_Updated_Construction_Pricing.pdf",
        "1wrfNB-k721-S_Xb_lecDcJpGJgFix1vp": "Freeman_proposal_and_SOV_7-15-26_Signed.pdf",
        "1hyVebJCoarbDocXJ8KftkecqX3NZg3Aj": "2025_Pricing_Freeman_Ranch.pdf",
    }
    # latest Arixa statements
    arixa_q = "name contains 'Arixa' and name contains 'Freeman' and trashed=false"
    arixa_files = (
        drive.files()
        .list(
            q=arixa_q,
            pageSize=20,
            fields="files(id,name,modifiedTime)",
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
            corpora="allDrives",
            orderBy="modifiedTime desc",
        )
        .execute()
        .get("files", [])
    )
    print("\n===== ARIXA FILES =====")
    for f in arixa_files:
        print(f"{f['modifiedTime']} | {f['id']} | {f['name']}")
        if "2026" in f["name"] or "Billing" in f["name"] or "Statement" in f["name"]:
            dest = OUT / ("".join(ch if ch.isalnum() or ch in "._- " else "_" for ch in f["name"])[:100])
            if not dest.exists():
                try:
                    download(drive, f["id"], dest)
                except Exception as e:
                    print("  download fail:", e)

    for fid, name in pdfs.items():
        dest = OUT / name
        if not dest.exists():
            download(drive, fid, dest)

    for p in sorted(OUT.glob("*.pdf")):
        dump_pdf_text(p, max_chars=5000)


if __name__ == "__main__":
    main()
